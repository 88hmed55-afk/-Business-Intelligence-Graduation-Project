from __future__ import annotations
import csv
import io
import os
from datetime import datetime
from typing import Any, Dict, List, Tuple

from app.core.constants import EXPORT_MAX_ROWS
from app.core.exceptions import BadRequestError
from app.shared.enums import ReportExportFormat
from app.shared.utils.helpers import safe_round

_CSV_CONTENT_TYPE = "text/csv"
_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_CONTENT_TYPE = "application/pdf"

_FONTS_DIR = os.path.join(os.path.dirname(__file__), "fonts")
_ARABIC_FONT_REGULAR = "Amiri-Regular.ttf"
_ARABIC_FONT_BOLD = "Amiri-Bold.ttf"

_REPORT_TITLES: Dict[str, Tuple[str, str]] = {
    "sales": ("Sales Report", "تقرير المبيعات"),
    "profit": ("Profit Report", "تقرير الأرباح"),
    "customers": ("Customers Report", "تقرير العملاء"),
    "products": ("Products Report", "تقرير المنتجات"),
    "inventory": ("Inventory Report", "تقرير المخزون"),
    "monthly": ("Monthly Report", "التقرير الشهري"),
    "yearly": ("Yearly Report", "التقرير السنوي"),
}


def _to_export_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime,)):
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _contains_arabic(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return any("\u0600" <= char <= "\u06FF" for char in value)


def _is_arabic_language(language: str | None) -> bool:
    if not language:
        return False
    return language.strip().lower().split(",")[0].split(";")[0].strip().startswith("ar")


def _shape_arabic(value: Any) -> Any:
    """Returns a bidi-shaped string suitable for an Arabic-capable font, or the original value."""
    if not _contains_arabic(value):
        return value
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        return get_display(arabic_reshaper.reshape(value))
    except Exception:
        return value


def _register_arabic_fonts() -> str | None:
    """Registers the bundled Arabic TTF fonts with reportlab. Returns the body font name or None."""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        regular_path = os.path.join(_FONTS_DIR, _ARABIC_FONT_REGULAR)
        bold_path = os.path.join(_FONTS_DIR, _ARABIC_FONT_BOLD)
        if not os.path.exists(regular_path):
            return None
        registered = pdfmetrics.getRegisteredFontNames()
        if "Amiri" not in registered:
            pdfmetrics.registerFont(TTFont("Amiri", regular_path))
        if "Amiri-Bold" not in registered and os.path.exists(bold_path):
            pdfmetrics.registerFont(TTFont("Amiri-Bold", bold_path))
        return "Amiri"
    except Exception:
        return None


class ExportService:
    """Serializes tabular data into CSV, XLSX or PDF byte streams."""

    def export(
        self,
        *,
        report_type: str,
        rows: List[Dict[str, Any]],
        format: ReportExportFormat,
        title: str,
        language: str | None = None,
    ) -> Tuple[str, bytes, str]:
        if len(rows) > EXPORT_MAX_ROWS:
            raise BadRequestError(f"Too many rows to export (max {EXPORT_MAX_ROWS}).")
        if not rows:
            columns: List[str] = []
        else:
            columns = list(rows[0].keys())

        handler = {
            ReportExportFormat.CSV: self._to_csv,
            ReportExportFormat.XLSX: self._to_xlsx,
            ReportExportFormat.PDF: self._to_pdf,
        }.get(format)
        if handler is None:
            raise BadRequestError("Unsupported export format.")

        content = handler(rows=rows, columns=columns, title=title, language=language)
        content_type = {
            ReportExportFormat.CSV: _CSV_CONTENT_TYPE,
            ReportExportFormat.XLSX: _XLSX_CONTENT_TYPE,
            ReportExportFormat.PDF: _PDF_CONTENT_TYPE,
        }[format]
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_{stamp}.{format.value}"
        return content_type, content, filename

    def _to_csv(
        self,
        *,
        rows: List[Dict[str, Any]],
        columns: List[str],
        title: str,
        language: str | None = None,
    ) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        if columns:
            writer.writerow(columns)
        for row in rows:
            writer.writerow([_to_export_value(row.get(col)) for col in columns])
        # utf-8-sig prepends a BOM so Excel opens Arabic text correctly.
        return buffer.getvalue().encode("utf-8-sig")

    def _to_xlsx(
        self,
        *,
        rows: List[Dict[str, Any]],
        columns: List[str],
        title: str,
        language: str | None = None,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = (title[:28] or "Report")

        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if columns:
            for col_idx, col in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
                cell.fill = header_fill
                cell.font = header_font
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, col in enumerate(columns, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=_to_export_value(row.get(col)))

        for col_idx in range(1, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 22

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _to_pdf(
        self,
        *,
        rows: List[Dict[str, Any]],
        columns: List[str],
        title: str,
        language: str | None = None,
    ) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        arabic_font = _register_arabic_fonts()
        # If the user asked for Arabic (or the data contains Arabic text) and we
        # have the font, render everything with Amiri and pre-shape Arabic runs.
        use_arabic = arabic_font is not None and (_is_arabic_language(language) or any(
            _contains_arabic(value) for row in rows for value in row.values()
        ))
        if use_arabic:
            body_font = arabic_font
            header_font_name = "Amiri-Bold" if _font_registered("Amiri-Bold") else arabic_font
            title = _shape_arabic(title)
        else:
            body_font = "Helvetica"
            header_font_name = "Helvetica-Bold"

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        if use_arabic:
            title_style.fontName = arabic_font
        elements = []
        elements.append(Paragraph(str(title), title_style))
        elements.append(Spacer(1, 12))

        header = [str(_shape_arabic(col.replace("_", " ").title())) for col in columns]
        table_data = [header]
        for row in rows:
            table_data.append([_shape_arabic(_to_export_value(row.get(col))) for col in columns])

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("FONTNAME", (0, 0), (-1, 0), header_font_name),
                    ("FONTNAME", (0, 1), (-1, -1), body_font),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()


def _font_registered(font_name: str) -> bool:
    try:
        from reportlab.pdfbase import pdfmetrics

        return font_name in pdfmetrics.getRegisteredFontNames()
    except Exception:
        return False
